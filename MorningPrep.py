# -*- coding: utf-8 -*-
"""
Created on Wed Oct 26 11:10:01 2022

@author: fben-eghan
"""

#Setup enivronment

import re
from os import remove
from shutil import move
import os
from win32com import client
import time
import pandas as pd
import numpy as np
from datetime import date
from shutil import copyfile, copy2
import glob

#STEP 1: saves the FOFs to X:\Portia

outlook = client.Dispatch("Outlook.Application")
inbox = outlook.GetNameSpace("MAPI").Folders['Data Support'].Folders['Inbox']
messages = inbox.Items


def save_attachments(subject_prefix): # changed parameter name
    messages.Sort("[ReceivedTime]", True)  # sort by received date: newest to oldest 
    for message in messages:
        if message.Subject.startswith(subject_prefix): # changed test
            print("saving attachments for:", message.Subject)
            for attachment in message.Attachments:
                print(attachment.FileName)
                attachment.SaveAsFile(r'X:\Portia\{}'.format(attachment.FileName))  # changed to file-name
                
            return  # exit after first matched message

save_attachments(r'Broadcasting Mail -- ASSETS')
save_attachments(r'Broadcasting Mail -- VALPOS')
save_attachments(r'Broadcasting Mail -- NFMGTR')
save_attachments(r'Broadcasting Mail -- TTRADE')


#STEP 2: Files the FOF Emails away 

def file_FOFS(subject_prefix): # changed parameter name
    messages.Sort("[ReceivedTime]", True)  # sort by received date: newest to oldest 
    for message in messages:
        if message.Subject.startswith(subject_prefix): # changed test
            print("filing:", message.Subject)
            message.UnRead = False
            message.Move(inbox.Folders['RBCD_FOFs'])
                
            return  # exit after first matched message

file_FOFS(r'Broadcasting Mail -- ASSETS')
file_FOFS(r'Broadcasting Mail -- VALPOS')
file_FOFS(r'Broadcasting Mail -- NFMGTR')
file_FOFS(r'Broadcasting Mail -- TTRADE')



def not_used(subject_prefix): # changed parameter name
    messages.Sort("[ReceivedTime]", True)  # sort by received date: newest to oldest 
    for message in messages:
        if message.Subject.startswith(subject_prefix): # changed test
            print("filing:", message.Subject)
            message.UnRead = False
            message.Move(inbox.Folders['RBCD_FOFs'].Folders['Not Used'])
            
            return  # exit after first matched message

not_used(r'Broadcasting Mail -- XC31P163VALPOS2')
not_used(r'Broadcasting Mail -- XC30P163REFTAB1')
not_used(r'Broadcasting Mail -- XC29P163CSHTRN1')
not_used(r'Broadcasting Mail -- NFNODV')
not_used(r'Broadcasting Mail -- CASHFC')
not_used(r'Broadcasting Mail -- NFMDIV')
not_used(r'Broadcasting Mail -- XC27P163FUNDEX1')
not_used(r'Broadcasting Mail -- XC28P163BROKER1')
not_used(r'Broadcasting Mail -- XC25P163EXRATE1')
not_used(r'Broadcasting Mail -- XC41P163REC571')
not_used(r'Broadcasting Mail -- XC40P163REC950')

#STEP 3: Preparation for running the rec
#Deletes Old rec files and adds new worksheets to Cash Rec Workbook

def RecPrep():
    import re
    from shutil import move
    from os import remove
    import os
    import time
    import pandas as pd
    import numpy as np
    from datetime import date
    from shutil import copyfile, copy2
    import datetime
    import shutil
    from glob import iglob
    from openpyxl import load_workbook
    
    
    lb_ = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%d%m%y')
    lb1 = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%Y%m%d')
    lb2 = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%Y%m%d')
    lb = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%m-%d-%y')
    lb3 = (pd.datetime.today() - pd.tseries.offsets.BDay(2)).date().strftime('%Y%m%d')
    y = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).year
    m = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%B')
    m_ = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%B')
    
    remove(r'Q:\Rec\Rec1.txt')
    remove(r'Q:\Rec\Cash.txt')

    src_folder = r'Q:\GNET\{}'.format(lb)
    pattern1 = src_folder + r'\\Cash1_{}_22*'.format(lb2)
    dst_folder = r'Q:\Rec'

        
    for file in iglob(pattern1, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        copy2(file, dst_folder)
        
    pattern2 = src_folder + r'\\Rec1_{}_22*'.format(lb2)

    for file in iglob(pattern2, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        copy2(file, dst_folder)
    
    
    patt_1 = dst_folder + r'\\Cash1_{}_22*'.format(lb2)       
      
    for file in iglob(patt_1, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        os.rename(file, dst_folder + r'\\Cash.txt')
        print('Copied and renamed:', file[-25:])
        
    patt_2 = dst_folder + r'\\Rec1_{}_22*'.format(lb2           
    
    for file in iglob(patt_2, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        os.rename(file, dst_folder + r'\\Rec1.txt')
        print('Copied and renamed:', file[-24:])       
        
    print("Ready to run Rec")
    
    
    def move_sheet(wb, from_loc=None, to_loc=None):
        sheets = wb._sheets
        
        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1
        #if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0
        
        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)   
    
    lb_1 = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%d%m%y')
    y = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).year

    path = r'Q:\Rec\Cash Rec {}.xlsx'.format(y)
    book = load_workbook(path)
    rec_diffs = r'Rec Diffs COB {}'.format(lb_1)
    cash_diffs = r'Cash Diffs COB {}'.format(lb_1)

    book.create_sheet(rec_diffs)
    move_sheet(book)
    book.create_sheet(cash_diffs)
    move_sheet(book)
    book.save(r'Q:\Rec\Cash Rec {}.xlsx'.format(y))
    print("Ready to update Cash Rec")
    
RecPrep()


def RecPrepUAT():
    import re
    from shutil import move
    from os import remove
    import os
    import time
    import pandas as pd
    import numpy as np
    from datetime import date
    from shutil import copyfile, copy2
    import datetime
    import shutil
    from glob import iglob
    from openpyxl import load_workbook
    
    
    lb_ = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%d%m%y')
    lb1 = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%Y%m%d')
    lb2 = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%Y%m%d')
    lb = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%m-%d-%y')
    lb3 = (pd.datetime.today() - pd.tseries.offsets.BDay(2)).date().strftime('%Y%m%d')
    y = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).year
    m = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%B')
    m_ = (pd.datetime.today() - pd.tseries.offsets.BDay(0)).date().strftime('%B')
    
    remove(r'Q:\RecNT\Rec1.txt')
    remove(r'Q:\RecNT\Cash.txt')

    src_folder = r'Q:\GNET\{}'.format(lb)
    pattern1 = src_folder + r'\\Cash2_{}_01*'.format(lb1)
    dst_folder = r'Q:\RecNT'

        
    for file in iglob(pattern1, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        copy2(file, dst_folder)
        
    pattern2 = src_folder + r'\\Rec2_{}_01*'.format(lb1)

    for file in iglob(pattern2, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        copy2(file, dst_folder)
    
    
    patt_1 = dst_folder + r'\\Cash2_{}_01*'.format(lb1)        
      
    for file in iglob(patt_1, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        os.rename(file, dst_folder + r'\\Cash.txt')
        print('Copied and renamed (NT):', file[-25:])
        
    patt_2 = dst_folder + r'\\Rec2_{}_01*'.format(lb1)          
    
    for file in iglob(patt_2, recursive=True):
        # extract file name form file path
        file_name = os.path.basename(file)
        os.rename(file, dst_folder + r'\\Rec1.txt')
        print('Copied and renamed (NT):', file[-24:])       
        
    print("Ready to run RecNT")
    
    
    def move_sheet(wb, from_loc=None, to_loc=None):
        sheets = wb._sheets
        
        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1
        #if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0
        
        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)   
    
    lb_1 = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).date().strftime('%d%m%y')
    y = (pd.datetime.today() - pd.tseries.offsets.BDay(1)).year

    path = r'Q:\RecNT\Cash Rec NT {}.xlsx'.format(y)
    book = load_workbook(path)
    rec_diffs = r'Rec Diffs COB {}'.format(lb_1)
    cash_diffs = r'Cash Diffs COB {}'.format(lb_1)

    book.create_sheet(rec_diffs)
    move_sheet(book)
    book.create_sheet(cash_diffs)
    move_sheet(book)
    book.save(r'Q:\RecNT\Cash Rec NT {}.xlsx'.format(y))
    print("Ready to update Cash Rec NT")
    
RecPrepUAT()

